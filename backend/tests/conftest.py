"""
Shared fixtures for all tests.
"""
import sys
from pathlib import Path

import pytest

# Add backend to sys.path so imports work
sys.path.insert(0, str(Path(__file__).parent.parent))


@pytest.fixture
def sample_csv_text():
    """Simple CSV content as string."""
    return "name,age,city\nAlice,30,Paris\nBob,25,Lyon\nCharlie,35,Marseille\n"


@pytest.fixture
def sample_tsv_text():
    """Tab-separated content."""
    return "name\tage\tcity\nAlice\t30\tParis\nBob\t25\tLyon\n"


@pytest.fixture
def sample_markdown_table():
    """Markdown table content."""
    return (
        "| Produit | Prix | Stock |\n"
        "|---------|------|-------|\n"
        "| Pommes  | 2.50 | 100   |\n"
        "| Bananes | 1.80 | 50    |\n"
        "| Oranges | 3.00 | 75    |\n"
    )


@pytest.fixture
def sample_xlsx_bytes():
    """Generate a minimal .xlsx file in memory."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventes"
    ws.append(["Produit", "Quantité", "Prix"])
    ws.append(["Pommes", 100, 2.50])
    ws.append(["Bananes", 50, 1.80])
    ws.append(["Oranges", 75, 3.00])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.fixture
def sample_xlsx_multisheet_bytes():
    """Generate a multi-sheet .xlsx file."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Ventes"
    ws1.append(["Produit", "Quantité"])
    ws1.append(["Pommes", 100])

    ws2 = wb.create_sheet("Achats")
    ws2.append(["Fournisseur", "Montant"])
    ws2.append(["FournisseurA", 500])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.fixture
def sample_bank_data():
    """Data that looks like bank transactions."""
    return [
        {"date": "2024-01-15", "description": "Carrefour", "montant": "-45.30", "solde": "1254.70"},
        {"date": "2024-01-16", "description": "SNCF Billet", "montant": "-89.00", "solde": "1165.70"},
        {"date": "2024-01-17", "description": "Salaire", "montant": "2500.00", "solde": "3665.70"},
    ]
