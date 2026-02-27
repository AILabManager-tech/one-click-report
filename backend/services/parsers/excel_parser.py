"""
Excel parser — supports .xlsx (openpyxl) and .xls (xlrd).
Auto-detects header row and handles multi-sheet workbooks.
"""
import io
import logging

from .base import BaseParser, ParseResponse

logger = logging.getLogger(__name__)

EXCEL_EXTENSIONS = {".xlsx", ".xls"}
EXCEL_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}
# Magic bytes: XLSX = PK (ZIP), XLS = D0 CF 11 E0 (OLE2)
XLSX_MAGIC = b"PK"
XLS_MAGIC = b"\xd0\xcf\x11\xe0"

MAX_ROWS = 10_000


class ExcelParser(BaseParser):
    def supports(self, filename: str, content_type: str) -> bool:
        ext = _get_ext(filename)
        return ext in EXCEL_EXTENSIONS or content_type in EXCEL_CONTENT_TYPES

    async def parse(
        self, content: bytes, filename: str, options: dict | None = None
    ) -> ParseResponse:
        options = options or {}
        _validate_magic(content, filename)

        ext = _get_ext(filename)
        if ext == ".xls":
            return _parse_xls(content, options)
        return _parse_xlsx(content, options)


def _get_ext(filename: str) -> str:
    filename = filename.lower()
    if filename.endswith(".xlsx"):
        return ".xlsx"
    if filename.endswith(".xls"):
        return ".xls"
    return ""


def _validate_magic(content: bytes, filename: str) -> None:
    """Validate file magic bytes to prevent disguised uploads."""
    ext = _get_ext(filename)
    if ext == ".xlsx" and not content[:2].startswith(XLSX_MAGIC):
        raise ValueError("Invalid XLSX file (bad magic bytes)")
    if ext == ".xls" and not content[:4].startswith(XLS_MAGIC):
        raise ValueError("Invalid XLS file (bad magic bytes)")


def _parse_xlsx(content: bytes, options: dict) -> ParseResponse:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    # Multi-sheet: ask user to choose
    selected_sheet = options.get("sheet")
    if len(sheet_names) > 1 and not selected_sheet:
        wb.close()
        return ParseResponse(
            input_type="excel",
            needs_user_input=True,
            options={"sheets": sheet_names},
            confidence=1.0,
            warnings=[],
        )

    ws = wb[selected_sheet] if selected_sheet and selected_sheet in sheet_names else wb.active
    data, columns = _extract_sheet_data(ws)
    wb.close()

    return ParseResponse(
        input_type="excel",
        data=data,
        rows=len(data),
        columns=columns,
        confidence=1.0,
        preview_rows=data[:10],
        warnings=[],
    )


def _parse_xls(content: bytes, options: dict) -> ParseResponse:
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    sheet_names = wb.sheet_names()

    selected_sheet = options.get("sheet")
    if len(sheet_names) > 1 and not selected_sheet:
        return ParseResponse(
            input_type="excel",
            needs_user_input=True,
            options={"sheets": sheet_names},
            confidence=1.0,
            warnings=[],
        )

    if selected_sheet and selected_sheet in sheet_names:
        ws = wb.sheet_by_name(selected_sheet)
    else:
        ws = wb.sheet_by_index(0)

    data, columns = _extract_xls_sheet_data(ws)

    return ParseResponse(
        input_type="excel",
        data=data,
        rows=len(data),
        columns=columns,
        confidence=1.0,
        preview_rows=data[:10],
        warnings=[],
    )


def _extract_sheet_data(ws) -> tuple[list[dict], list[str]]:
    """Extract data from an openpyxl worksheet. Auto-detects header row."""
    rows_raw = []
    for row in ws.iter_rows(values_only=True):
        rows_raw.append(list(row))
        if len(rows_raw) > MAX_ROWS + 5:
            break

    if not rows_raw:
        return [], []

    header_idx = _detect_header_row(rows_raw)
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows_raw[header_idx])]

    data = []
    for row in rows_raw[header_idx + 1 : header_idx + 1 + MAX_ROWS]:
        if all(c is None for c in row):
            continue
        record = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            record[key] = _clean_value(val)
        data.append(record)

    return data, headers


def _extract_xls_sheet_data(ws) -> tuple[list[dict], list[str]]:
    """Extract data from an xlrd worksheet."""
    if ws.nrows == 0:
        return [], []

    rows_raw = []
    for rx in range(min(ws.nrows, MAX_ROWS + 5)):
        rows_raw.append(ws.row_values(rx))

    header_idx = _detect_header_row(rows_raw)
    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(rows_raw[header_idx])]

    data = []
    for row in rows_raw[header_idx + 1 : header_idx + 1 + MAX_ROWS]:
        if all(c is None or c == "" for c in row):
            continue
        record = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            record[key] = _clean_value(val)
        data.append(record)

    return data, headers


def _detect_header_row(rows: list[list], max_check: int = 5) -> int:
    """Find header row: first row where >50% of cells are non-empty strings."""
    for idx in range(min(max_check, len(rows))):
        row = rows[idx]
        if not row:
            continue
        str_count = sum(1 for c in row if isinstance(c, str) and c.strip())
        non_empty = sum(1 for c in row if c is not None and c != "")
        if non_empty > 0 and str_count / max(non_empty, 1) > 0.5:
            return idx
    return 0


def _clean_value(val):
    """Convert Excel cell values to JSON-safe types."""
    if val is None:
        return None
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val
